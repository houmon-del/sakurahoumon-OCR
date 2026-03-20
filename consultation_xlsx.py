# -*- coding: utf-8 -*-
"""相談シートExcelテンプレートへのデータ書き込み"""

import io
import os
import re
import openpyxl
from openpyxl.styles import Alignment

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "static", "templates", "consultation_template.xlsx")

# 来院理由マッピング（テキストセルの座標 ← 修正済み）
VISIT_REASON_CELLS = {
    "歯が痛い":               "C45",   # C45:I45
    "グラグラ":               "K45",   # K45:Q45
    "歯ぐきが腫れた・痛い":   "S45",   # S45:Y45
    "つめもの・かぶせものが取れた": "AA45",  # AA45:AG45
    "口の中にできものがある":  "C46",   # C46:I46
    "入れ歯の調子が悪い":     "K46",   # K46:Q46
    "入れ歯を作りたい":       "S46",   # S46:Y46
    "歯が抜けた":             "AA46",  # AA46:AG46
    "口腔ケアをして欲しい":   "C47",   # C47:I47
    "その他":                 "N47",   # N47:AF47 (入力欄)
}

# 要介護度マッピング（構造化データの値 → テンプレート内の選択肢テキスト）
CARE_LEVEL_MAP = {
    "要支援1": "支1", "要支援2": "支2",
    "要介護1": "介1", "要介護2": "介2", "要介護3": "介3",
    "要介護4": "介４", "要介護5": "介５",
}

# 食事形態チェックボックスセル（□→■）
DIET_CHECKBOX = {
    "経口摂取": "F22",
    "常食":     "K22",
    "嚥下調整食": "N22",
    # 嚥下レベル (嚥下調整食の後に続く数字コード)
    "4":   "S22", "3":   "U22",
    "2-2": "W22", "2-1": "Y22",
    "1j":  "AA22", "0t": "AC22", "0j": "AE22",
    "経腸栄養":   "F23",
    "静脈栄養":   "K23",
}
ASPIRATION_CHECKBOX = {
    "あり": "F24",
    "なし": "AA24",
}

# 知ったきっかけマッピング
REFERRAL_SOURCE_CELLS = {
    "HP":           "B52",
    "口コミ":        "B52",
    "院内パンフレット": "B52",
    "スタッフから":   "B52",
    "ポスターを見て": "B52",
    "他患者もさくら会に依頼しているため": "B52",
    "広告物":        "B53",
}

# 訪問曜日セルマッピング {('am'/'pm', '日'..'土'): cell}
SCHEDULE_CELLS = {
    ("am", "日"): "F26", ("pm", "日"): "F28",
    ("am", "月"): "J26", ("pm", "月"): "J28",
    ("am", "火"): "N26", ("pm", "火"): "N28",
    ("am", "水"): "R26", ("pm", "水"): "R28",
    ("am", "木"): "V26", ("pm", "木"): "V28",
    ("am", "金"): "Z26", ("pm", "金"): "Z28",
    ("am", "土"): "AD26", ("pm", "土"): "AD28",
}


def _top_left(ws, cell_addr):
    """マージ範囲の左上セルアドレスを返す。非マージなら cell_addr そのまま"""
    for mr in ws.merged_cells.ranges:
        if cell_addr in mr:
            col = openpyxl.utils.get_column_letter(mr.min_col)
            return f"{col}{mr.min_row}"
    return cell_addr


def _mark_checkbox(ws, cell_addr):
    """□ → ■ に変更（チェックボックスをONに）"""
    cell = ws[cell_addr]
    if cell.value == "□":
        cell.value = "■"
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _w(ws, cell_addr, value):
    """空白セル・数式セルへの通常書き込み"""
    addr = _top_left(ws, cell_addr)
    ws[addr] = value
    ws[addr].alignment = Alignment(wrap_text=True, vertical="center")


def _mark_option(ws, cell_addr, value):
    """
    既存の選択肢テキスト内で value を ● でマーク。
    テキストなし・数式セルの場合は普通に書く。
    value が選択肢に見つからない場合は末尾に追記。
    既存テキストのあるセルはアライメントを保持（行高さ崩壊防止）。
    """
    if not value:
        return
    addr = _top_left(ws, cell_addr)
    current = ws[addr].value

    # 空セル・数式セル → 通常書き込み（アライメント設定OK）
    if current is None:
        ws[addr] = value
        ws[addr].alignment = Alignment(wrap_text=True, vertical="center")
        return
    current_str = str(current)
    if current_str.startswith("="):
        ws[addr] = value
        ws[addr].alignment = Alignment(wrap_text=True, vertical="center")
        return

    value_str = str(value)
    # 検索候補: そのまま → 末尾の接尾辞を除いた形 → 先頭1文字（元号用）
    candidates = [value_str]
    for suffix in ["性", "割", "和", "成", "正", "治", "ジャー"]:
        if value_str.endswith(suffix) and len(value_str) > 1:
            candidates.append(value_str[:-len(suffix)])
    if len(value_str) > 1:
        candidates.append(value_str[0])

    for cand in candidates:
        if cand and cand in current_str and "●" + cand not in current_str:
            ws[addr] = current_str.replace(cand, "●" + cand, 1)
            # 既存テキストのセルはアライメントを変更しない（行高さ維持）
            return

    # 見つからない場合は末尾に追記（アライメント変更なし）
    ws[addr] = current_str.rstrip() + " (" + value_str + ")"


def _fill_parens(ws, cell_addr, value):
    """
    既存テキストの（　　...）または(　　...)内にvalueを挿入。
    括弧内スペースが見つからない場合は通常書き込み。
    """
    if not value:
        return
    addr = _top_left(ws, cell_addr)
    current = ws[addr].value

    if current is None:
        ws[addr] = value
        ws[addr].alignment = Alignment(wrap_text=True, vertical="center")
        return
    current_str = str(current)
    if current_str.startswith("="):
        ws[addr] = value
        ws[addr].alignment = Alignment(wrap_text=True, vertical="center")
        return

    # 全角・半角の丸括弧内のスペース（半角・全角）をvalueで置換
    new_val = re.sub(r'[（(][\s\u3000]+[）)]', f'（{value}）', current_str, count=1)
    if new_val != current_str:
        ws[addr] = new_val
    else:
        ws[addr] = value
    # 既存テキストのセルはアライメントを変更しない（行高さ維持）


def fill_template(structured: dict) -> bytes:
    """構造化データをテンプレートに埋め込みxlsxバイト列を返す"""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    p     = structured.get("patient", {})
    c     = structured.get("contact", {})
    ins   = structured.get("insurance", {})
    mh    = structured.get("medical_history", {})
    inf   = structured.get("infection", {})
    phy   = structured.get("physician", {})
    diet  = structured.get("diet", {})
    sched = structured.get("schedule", {})
    req   = structured.get("requester", {})
    kp    = structured.get("key_person", {})
    cm    = structured.get("care_manager", {})

    # ── 本人情報 ──
    furigana = f"{p.get('furigana_sei','') or ''} {p.get('furigana_mei','') or ''}".strip()
    _w(ws, "F5", furigana)   # F5:N5 は =PHONETIC 数式 → エクスポート用に上書きOK

    name = f"{p.get('sei','') or ''} {p.get('mei','') or ''}".strip()
    _w(ws, "F6", name)

    # 性別: 選択値のみ書き込み
    if p.get("gender"):
        _w(ws, "O5", p["gender"])

    # 生年月日
    dob_era  = p.get("dob_era", "")
    dob_year = p.get("dob_year", "")
    dob_mon  = p.get("dob_month", "")
    dob_day  = p.get("dob_day", "")
    if dob_era:  _w(ws, "U5", dob_era)  # 元号: 選択値のみ
    if dob_year: _w(ws, "Y5",  dob_year)
    if dob_mon:  _w(ws, "AB5", dob_mon)
    if dob_day:  _w(ws, "AE5", dob_day)

    # 年齢: Y7:AG7
    if p.get("age"):
        _w(ws, "Y7", f'{p["age"]}歳')

    # 郵便番号: G8（〒ラベルの右）に番号のみ
    if p.get("postal_code"):
        _w(ws, "G8", p["postal_code"])

    # 施設名・住所・部屋番号: F9:U10 に統合
    addr_parts = []
    if p.get("facility"): addr_parts.append(p["facility"])
    if p.get("address"):  addr_parts.append(p["address"])
    room = p.get("room", "")
    if room: addr_parts.append(f"部屋番号（{room}）")
    if addr_parts:
        ws["F9"] = "\n".join(addr_parts)
        ws["F9"].alignment = Alignment(wrap_text=True, vertical="top")

    # 駐車場: 選択値のみ書き込み
    if p.get("parking"):
        _w(ws, "Z8", p["parking"])

    # ── 電話番号 ──
    if c.get("home_phone"):   _w(ws, "H11", c["home_phone"])
    if c.get("mobile_phone"): _w(ws, "H13", c["mobile_phone"])

    # ── 保険情報 ──
    # 医療負担割合: 選択値のみ書き込み
    if ins.get("burden_ratio"):
        ratio = str(ins["burden_ratio"]).replace("割", "")
        _w(ws, "R11", f"{ratio}割")
    # 公費: 括弧内に挿入
    if ins.get("public_expense"):
        _fill_parens(ws, "Z11", ins["public_expense"])
    # 要介護度: 選択値のみ書き込み
    if ins.get("care_level"):
        _w(ws, "Z13", ins["care_level"])

    # ── 既往歴: 選択された病名のみ書き込み ──
    conditions = mh.get("conditions") or []
    cond_parts = list(conditions)
    if mh.get("other"): cond_parts.append("その他: " + mh["other"])
    if cond_parts:
        _w(ws, "F15", "　".join(cond_parts))

    # ── 感染症: 選択値のみ書き込み ──
    inf_status  = inf.get("status", "")
    inf_details = inf.get("details", "")
    if isinstance(inf_details, list): inf_details = "、".join(inf_details)
    inf_text = inf_status or ""
    if inf_details: inf_text += "（" + inf_details + "）"
    if inf_text:
        _w(ws, "F18", inf_text)

    # ── 内科主治医 ──
    if phy.get("hospital"): _w(ws, "I19", phy["hospital"])   # I19:R19 空白入力欄
    if phy.get("doctor"):   _w(ws, "W19", phy["doctor"])     # W19:AF19 空白入力欄

    # ── 意思疎通: 選択値のみ書き込み ──
    if structured.get("communication"):
        _w(ws, "F21", structured["communication"])

    # ── 食事形態: □チェックボックスをマーク ──
    diet_type = diet.get("type", "") or ""
    if "経口摂取" in diet_type:
        _mark_checkbox(ws, DIET_CHECKBOX["経口摂取"])
        if "常食" in diet_type:
            _mark_checkbox(ws, DIET_CHECKBOX["常食"])
        if "嚥下調整食" in diet_type:
            _mark_checkbox(ws, DIET_CHECKBOX["嚥下調整食"])
            # 嚥下レベル (例: "レベル3", "コード2-1" など)
            for level_code, level_cell in [
                ("4", "S22"), ("3", "U22"), ("2-2", "W22"), ("2-1", "Y22"),
                ("1j", "AA22"), ("0t", "AC22"), ("0j", "AE22"),
            ]:
                if level_code in diet_type:
                    _mark_checkbox(ws, level_cell)
                    break
    elif "経腸栄養" in diet_type:
        _mark_checkbox(ws, DIET_CHECKBOX["経腸栄養"])
    elif "静脈栄養" in diet_type:
        _mark_checkbox(ws, DIET_CHECKBOX["静脈栄養"])

    # 誤嚥性肺炎
    aspiration = diet.get("aspiration_pneumonia", "") or ""
    if "あり" in aspiration:
        _mark_checkbox(ws, ASPIRATION_CHECKBOX["あり"])
    elif "なし" in aspiration:
        _mark_checkbox(ws, ASPIRATION_CHECKBOX["なし"])

    # ── 訪問可能曜日 ──
    days = ["日", "月", "火", "水", "木", "金", "土"]
    for slot in ("am", "pm"):
        slot_data = sched.get(slot) or {}
        for day in days:
            val = slot_data.get(day, "")
            cell_addr = SCHEDULE_CELLS.get((slot, day))
            if cell_addr and val:
                _w(ws, cell_addr, val)

    # ── 依頼者: 選択値のみ書き込み ──
    if req.get("type"):
        _w(ws, "G32", req["type"])
    req_parts = [x for x in [req.get("name"), req.get("phone")] if x]
    if req_parts: _w(ws, "V32", "　".join(req_parts))

    # ── キーパーソン ──
    kp_furi  = kp.get("furigana", "")
    kp_name  = kp.get("name", "")
    kp_rel   = kp.get("relationship", "")
    kp_phone = kp.get("phone", "")
    kp_addr  = kp.get("address", "")
    if kp_furi:  _w(ws, "J33", kp_furi)
    if kp_rel:   _w(ws, "W33", kp_rel)
    if kp_name:  _w(ws, "J34", kp_name)
    if kp_phone: _w(ws, "J35", kp_phone)
    if kp_addr:  _w(ws, "G36", kp_addr)

    # ── ケアマネジャー ──
    if cm.get("furigana"): _w(ws, "F40", cm["furigana"])
    if cm.get("phone"):    _w(ws, "W40", cm["phone"])
    if cm.get("name"):     _w(ws, "F41", cm["name"])
    if cm.get("facility"): _w(ws, "F42", cm["facility"])
    if cm.get("fax"):      _w(ws, "W42", cm["fax"])

    # ── 来院理由: テキストセルの先頭に ● を追加 ──
    visit_reasons = structured.get("visit_reason") or []
    for reason in visit_reasons:
        for key, cell_addr in VISIT_REASON_CELLS.items():
            if key in reason:
                addr = _top_left(ws, cell_addr)
                current = str(ws[addr].value or "")
                if not current.startswith("●"):
                    ws[addr] = "●" + current
                    # アライメント変更なし（行高さ維持）
                break

    # ── 備考 ──
    if structured.get("notes"):
        _w(ws, "E48", structured["notes"])

    # ── 知ったきっかけ: 選択値のみ書き込み ──
    referral_sources = structured.get("referral_source") or []
    b52_selected = []
    b53_selected = []
    b54_selected = []
    for source in referral_sources:
        matched = False
        for key, cell_addr in REFERRAL_SOURCE_CELLS.items():
            if key in source:
                if cell_addr == "B52":
                    b52_selected.append(key)
                elif cell_addr == "B53":
                    b53_selected.append(source)
                matched = True
                break
        if not matched:
            b54_selected.append(source)
    if b52_selected: _w(ws, "B52", "　".join(b52_selected))
    if b53_selected: _fill_parens(ws, "B53", "　".join(b53_selected))
    if b54_selected: _fill_parens(ws, "E54", "　".join(b54_selected))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
