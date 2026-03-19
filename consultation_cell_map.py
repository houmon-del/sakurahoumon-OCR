# -*- coding: utf-8 -*-
"""
相談シートテンプレートのセル座標マップを生成する。

Excelの列幅・行高からピクセル座標を計算し、
各データフィールドのマージセル範囲をJSONとして保存する。

使い方:
    python consultation_cell_map.py  → cell_map.json を生成
"""

import json
import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "static", "templates", "consultation_template.xlsx")
OUTPUT_PATH   = os.path.join(os.path.dirname(__file__), "static", "templates", "cell_map.json")

# 基準DPI（この値でピクセル座標を計算。スキャン時のDPIに合わせてスケール変換する）
BASE_DPI = 150

# ──────────────────────────────────────────────
# データフィールド定義
# (field_path, write_cell)
# field_pathはstructured dictのパスに対応
# ──────────────────────────────────────────────
DATA_FIELDS = [
    ("patient.furigana",         "F5"),
    ("patient.name",             "F6"),
    ("patient.gender",           "O5"),
    ("patient.dob_era",          "U5"),
    ("patient.dob_year",         "AB5"),
    ("patient.dob_month",        "AE5"),
    ("patient.dob_day",          "AH5"),
    ("patient.age",              "Y7"),
    ("patient.address",          "G8"),
    ("patient.room",             "G9"),
    ("contact.home_phone",       "H11"),
    ("contact.mobile_phone",     "H13"),
    ("insurance.burden_ratio",   "S11"),
    ("insurance.public_expense", "W11"),
    ("insurance.care_level",     "S13"),
    ("medical_history.text",     "F15"),
    ("infection.text",           "F18"),
    ("physician.hospital",       "G19"),
    ("physician.doctor",         "U19"),
    ("communication",            "F21"),
    ("diet.type",                "H22"),
    # 訪問曜日 AM/PM × 日〜土 は schedule_cells で別定義
    ("requester.type",           "G32"),
    ("requester.name_phone",     "V32"),
    ("key_person.furigana",      "J33"),
    ("key_person.relationship",  "W33"),
    ("key_person.name",          "J34"),
    ("key_person.phone",         "G35"),
    ("key_person.address",       "B36"),
    ("care_manager.furigana",    "F40"),
    ("care_manager.phone",       "W40"),
    ("care_manager.name",        "F41"),
    ("care_manager.facility",    "C42"),
    ("care_manager.fax",         "W42"),
    ("notes",                    "C48"),
    ("referral_source",          "C52"),
]

# 訪問曜日セル (slot, day) → cell
SCHEDULE_CELLS = {
    ("am", "日"): "F26",  ("pm", "日"): "F28",
    ("am", "月"): "J26",  ("pm", "月"): "J28",
    ("am", "火"): "N26",  ("pm", "火"): "N28",
    ("am", "水"): "R26",  ("pm", "水"): "R28",
    ("am", "木"): "V26",  ("pm", "木"): "V28",
    ("am", "金"): "Z26",  ("pm", "金"): "Z28",
    ("am", "土"): "AD26", ("pm", "土"): "AD28",
}

# アンカーラベル（位置合わせ用の印刷済みテキスト）
# OCRが読み取るはずの固定テキストと、その代表セル
ANCHOR_LABELS = [
    ("ふりがな", "B5"),   # 患者ふりがなラベル
    ("名前",     "B6"),
    ("住所",     "B8"),
    ("電話番号", "B11"),
    ("既往歴",   "B15"),
    ("感染症",   "B18"),
    ("内科主治医","B19"),
    ("依頼者",   "B32"),
]

# ──────────────────────────────────────────────
# 実測キャリブレーション済みセル座標
# （実際のA4スキャン @ 300dpi で観測した相対座標）
# Excelのメタデータでは印刷スケールが合わないため、
# 実スキャン上の位置を直接定義する。
# ──────────────────────────────────────────────
CALIBRATED_FIELDS = [
    # field_path,               [x1_rel, y1_rel, x2_rel, y2_rel]
    ("patient.furigana",         [0.15, 0.135, 0.41, 0.170]),
    ("patient.name",             [0.15, 0.160, 0.41, 0.200]),
    ("patient.gender",           [0.44, 0.140, 0.47, 0.205]),
    ("patient.dob_era",          [0.55, 0.140, 0.70, 0.180]),
    ("patient.dob_year",         [0.68, 0.140, 0.88, 0.175]),
    ("patient.dob_month",        [0.83, 0.140, 0.94, 0.175]),
    ("patient.dob_day",          [0.92, 0.140, 0.99, 0.175]),
    ("patient.age",              [0.70, 0.165, 0.99, 0.205]),
    ("patient.address",          [0.15, 0.193, 0.58, 0.228]),
    ("patient.room",             [0.35, 0.222, 0.62, 0.255]),
    ("contact.home_phone",       [0.23, 0.250, 0.50, 0.285]),
    ("contact.mobile_phone",     [0.23, 0.272, 0.50, 0.308]),
    ("insurance.burden_ratio",   [0.48, 0.263, 0.62, 0.305]),
    ("insurance.public_expense", [0.62, 0.250, 0.76, 0.295]),
    ("insurance.care_level",     [0.68, 0.268, 0.90, 0.305]),
    ("medical_history.text",     [0.15, 0.295, 0.98, 0.370]),
    ("infection.text",           [0.15, 0.350, 0.80, 0.390]),
    ("physician.hospital",       [0.15, 0.368, 0.72, 0.408]),
    ("physician.doctor",         [0.60, 0.368, 0.92, 0.408]),
    ("communication",            [0.15, 0.395, 0.98, 0.430]),
    ("diet.type",                [0.15, 0.415, 0.55, 0.450]),
    ("requester.type",           [0.15, 0.560, 0.55, 0.592]),
    ("requester.name_phone",     [0.55, 0.560, 0.98, 0.592]),
    ("key_person.furigana",      [0.27, 0.578, 0.52, 0.615]),
    ("key_person.relationship",  [0.60, 0.578, 0.92, 0.638]),
    ("key_person.name",          [0.27, 0.595, 0.52, 0.632]),
    ("key_person.phone",         [0.60, 0.617, 0.92, 0.653]),
    ("key_person.address",       [0.15, 0.635, 0.55, 0.675]),
    ("care_manager.furigana",    [0.15, 0.683, 0.52, 0.718]),
    ("care_manager.phone",       [0.60, 0.683, 0.90, 0.720]),
    ("care_manager.name",        [0.15, 0.710, 0.52, 0.732]),
    ("care_manager.facility",    [0.15, 0.730, 0.52, 0.758]),
    ("care_manager.fax",         [0.60, 0.718, 0.90, 0.758]),
    ("notes",                    [0.07, 0.818, 0.65, 0.858]),
    ("referral_source",          [0.07, 0.858, 0.78, 0.900]),
]

# スケジュールセルのキャリブレーション済み座標
# AM行: y≈0.497-0.524, PM行: y≈0.518-0.547
# 日〜土の各列X座標（各dayの中心x_relから±0.025）
CALIBRATED_SCHEDULE = []
_day_x = {"日": 0.210, "月": 0.325, "火": 0.427, "水": 0.531,
          "木": 0.638, "金": 0.733, "土": 0.844}
for _day, _cx in _day_x.items():
    CALIBRATED_SCHEDULE.append({
        "slot": "am", "day": _day,
        "cell": SCHEDULE_CELLS[("am", _day)],
        "rel": [_cx-0.028, 0.492, _cx+0.028, 0.527],
    })
    CALIBRATED_SCHEDULE.append({
        "slot": "pm", "day": _day,
        "cell": SCHEDULE_CELLS[("pm", _day)],
        "rel": [_cx-0.028, 0.515, _cx+0.028, 0.550],
    })


def col_width_to_px(w, dpi=BASE_DPI):
    if w is None or w == 0:
        w = 8.43  # Excel デフォルト幅
    return max(1, int((w * 7 + 5) * dpi / 96))


def row_height_to_px(h, dpi=BASE_DPI):
    if h is None or h == 0:
        h = 15.0  # Excel デフォルト高さ
    return max(1, int(h * dpi / 72))


def build_coord_tables(ws, dpi=BASE_DPI):
    """各列・行のピクセル開始座標テーブルを返す（0-indexed, 1-basedセルに対応）"""
    max_col = ws.max_column
    max_row = ws.max_row

    col_x = [0] * (max_col + 2)
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        w = ws.column_dimensions[letter].width or 8.43
        col_x[c] = col_x[c - 1] + col_width_to_px(w, dpi)
    col_x[max_col + 1] = col_x[max_col]  # 右端

    row_y = [0] * (max_row + 2)
    for r in range(1, max_row + 1):
        h = ws.row_dimensions[r].height or 15.0
        row_y[r] = row_y[r - 1] + row_height_to_px(h, dpi)
    row_y[max_row + 1] = row_y[max_row]

    return col_x, row_y


def get_merge_extent(ws, addr):
    """セルアドレスが属するマージ範囲の (min_col, min_row, max_col, max_row) を返す"""
    for mr in ws.merged_cells.ranges:
        if addr in mr:
            return mr.min_col, mr.min_row, mr.max_col, mr.max_row
    col = column_index_from_string("".join(c for c in addr if c.isalpha()))
    row = int("".join(c for c in addr if c.isdigit()))
    return col, row, col, row


def cell_to_bbox(col_x, row_y, min_col, min_row, max_col, max_row):
    """マージ範囲 → ピクセルbbox [x1, y1, x2, y2]"""
    return [col_x[min_col - 1], row_y[min_row - 1],
            col_x[max_col],     row_y[max_row]]


def generate():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    col_x, row_y = build_coord_tables(ws)

    total_w = col_x[ws.max_column]
    total_h = row_y[ws.max_row]

    def to_rel(bbox):
        """ピクセルbbox → 相対座標 [0-1] に変換（印刷スケールに依存しない）"""
        x1, y1, x2, y2 = bbox
        return [x1/total_w, y1/total_h, x2/total_w, y2/total_h]

    # データフィールドのbbox — キャリブレーション済み座標を優先使用
    calib_map = {f: rel for f, rel in CALIBRATED_FIELDS}
    fields = []
    for field_path, cell in DATA_FIELDS:
        if field_path in calib_map:
            rel = calib_map[field_path]
        else:
            extent = get_merge_extent(ws, cell)
            bbox   = cell_to_bbox(col_x, row_y, *extent)
            rel    = to_rel(bbox)
        fields.append({
            "field": field_path,
            "cell":  cell,
            "rel":   rel,
        })

    # スケジュールセルのbbox — キャリブレーション済み
    schedule = list(CALIBRATED_SCHEDULE)

    # アンカーラベルのbbox（相対座標）
    anchors = []
    for label_text, cell in ANCHOR_LABELS:
        extent = get_merge_extent(ws, cell)
        bbox   = cell_to_bbox(col_x, row_y, *extent)
        anchors.append({
            "text": label_text,
            "cell": cell,
            "rel":  to_rel(bbox),
        })

    cell_map = {
        "base_dpi":   BASE_DPI,
        "template_w": total_w,
        "template_h": total_h,
        "fields":     fields,
        "schedule":   schedule,
        "anchors":    anchors,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(cell_map, f, ensure_ascii=False, indent=2)

    print(f"生成完了: {OUTPUT_PATH}")
    print(f"テンプレートサイズ: {total_w} x {total_h} px @ {BASE_DPI}dpi")
    print(f"データフィールド数: {len(fields)}")
    print(f"スケジュールセル数: {len(schedule)}")
    return cell_map


if __name__ == "__main__":
    generate()
